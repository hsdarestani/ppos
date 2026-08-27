import csv
import io
import json
import re
import secrets
import sqlite3
from datetime import datetime
from functools import wraps
from itertools import chain

from flask import flash, redirect, request, session, url_for
from openpyxl import load_workbook

from verticals import VERTICALS

PERSIAN_DIGITS = str.maketrans('۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩', '01234567890123456789')

HEADER_ALIASES = {
    'business_name': [
        'business_name', 'name', 'title', 'place_name',
        'نام کسب و کار', 'نام کسب‌وکار', 'نام فروشگاه', 'نام مرکز',
        'نام فروشگاه یا مرکز', 'نام واحد', 'واحد صنفی', 'نام واحد صنفی',
    ],
    'mobile': [
        'mobile', 'mobile_phone', 'cellphone', 'cell_phone',
        'موبایل', 'موبايل', 'شماره موبایل', 'شماره موبايل', 'شماره همراه', 'تلفن همراه',
    ],
    'phone_any': [
        'phone', 'phone_number', 'phonenumber', 'phone number', 'contact_phone',
        'formatted_phone_number', 'international_phone_number', 'شماره تماس',
    ],
    'landline': ['landline', 'fixed_phone', 'telephone', 'tel', 'تلفن', 'تلفن ثابت', 'شماره ثابت'],
    'city': ['city', 'municipality', 'locality', 'شهر', 'شهرستان'],
    'address': ['address', 'full_address', 'formatted_address', 'street_address', 'آدرس', 'ادرس', 'آدرس پستی', 'نشانی'],
    'category': [
        'category', 'categories', 'type', 'types', 'subtype', 'subtypes', 'guild', 'vertical',
        'صنف', 'نوع صنف', 'تفکیک صنف', 'رسته', 'دسته بندی', 'دسته‌بندی', 'گروه شغلی',
    ],
    'owner': ['owner', 'owner_name', 'نام مسئول', 'نام فرد مسئول', 'مدیر', 'نام مدیر'],
    'instagram': ['instagram', 'instagram_url', 'social links', 'social_links', 'اینستاگرام', 'اینستا'],
    'logo_url': ['logo_url', 'logo', 'image', 'images', 'photo', 'thumbnail', 'لوگو'],
    'website': ['website', 'site', 'web_site', 'domain', 'وبسایت', 'وب سایت', 'سایت'],
    'maps_url': ['google_maps_url', 'google_url', 'maps_url', 'place_url', 'google_maps_link', 'map_link', 'link'],
    'place_id': ['place_id', 'google_place_id', 'google_id', 'google_maps_id', 'cid'],
    'rating': ['rating', 'stars', 'score', 'امتیاز'],
    'reviews': [
        'reviews', 'reviews count', 'reviewscount', 'reviews_count', 'review_count',
        'user_ratings_total', 'number_of_reviews', 'تعداد نظر', 'تعداد نظرات',
    ],
    'latitude': ['latitude', 'location lat', 'location_lat', 'lat', 'عرض جغرافیایی'],
    'longitude': ['longitude', 'location lng', 'location_lng', 'lng', 'lon', 'long', 'طول جغرافیایی'],
    'opening_hours': ['opening_hours', 'working hours', 'working_hours', 'hours', 'business_hours', 'ساعات کاری'],
    'status': ['business_status', 'status', 'وضعیت'],
}

CATEGORY_KEYWORDS = {
    'realestate': ['املاک', 'مشاور املاک', 'بنگاه', 'مسکن', 'real estate', 'real estate agency', 'property'],
    'beauty': ['آرایش زنانه', 'آرایشگاه زنانه', 'سالن زیبایی', 'زیبایی زنانه', 'beauty salon', 'hair salon', 'nail salon'],
    'barber': ['آرایش مردانه', 'آرایشگاه مردانه', 'پیرایش مردانه', 'barber', 'barber shop'],
    'auto': ['اتوگالری', 'نمایشگاه اتومبیل', 'نمایشگاه خودرو', 'خرید و فروش خودرو', 'car dealer', 'used car dealer'],
    'aesthetic': ['کلینیک زیبایی', 'مرکز زیبایی', 'پوست و مو', 'aesthetic clinic', 'skin care clinic', 'laser hair removal service'],
    'dentist': ['دندانپزشک', 'دندانپزشکی', 'کلینیک دندان', 'dentist', 'dental clinic'],
    'gym': ['باشگاه بدنسازی', 'باشگاه ورزشی', 'فیتنس', 'بدنسازی', 'gym', 'fitness center', 'pilates studio'],
    'trainer': ['مربی خصوصی', 'مربی شخصی', 'پرسونال ترینر', 'personal trainer'],
    'language': ['آموزشگاه زبان', 'موسسه زبان', 'language school', 'english language school'],
    'education': ['آموزشگاه', 'کنکور', 'مشاوره تحصیلی', 'education center', 'training center', 'tutoring service'],
    'repair': ['تعمیرگاه خودرو', 'مکانیکی', 'اتو سرویس', 'تعمیر اتومبیل', 'auto repair shop', 'car repair', 'mechanic'],
    'parts': ['لوازم یدکی', 'قطعات خودرو', 'یدکی اتومبیل', 'auto parts store', 'car parts'],
    'carwash': ['کارواش', 'دیتیلینگ', 'صفرشویی', 'car wash', 'car detailing service'],
    'fashion': ['مزون', 'لباس مجلسی', 'پوشاک زنانه', 'clothing store', 'dress store', 'boutique'],
    'gold': ['طلافروشی', 'طلا و جواهر', 'جواهرفروشی', 'jewelry store', 'jeweler', 'gold dealer'],
    'furniture': ['مبلمان', 'مبل فروشی', 'فروش مبلمان', 'furniture store'],
    'cabinet': ['کابینت', 'کابینت سازی', 'کابینت‌سازی', 'cabinet maker', 'kitchen remodeler'],
    'restaurant': ['رستوران', 'کافه', 'کافی شاپ', 'فست فود', 'restaurant', 'cafe', 'coffee shop', 'fast food restaurant'],
    'pet': ['پت شاپ', 'پت‌شاپ', 'دامپزشکی', 'کلینیک دامپزشکی', 'pet store', 'veterinarian', 'animal hospital'],
    'mobile': ['موبایل فروشی', 'فروش موبایل', 'گوشی موبایل', 'cell phone store', 'mobile phone shop'],
    'immigration': ['مهاجرت', 'موسسه مهاجرتی', 'خدمات مهاجرتی', 'immigration consultant', 'visa consultant'],
    'travel': ['آژانس مسافرتی', 'خدمات مسافرتی', 'تور و گردشگری', 'travel agency', 'tour agency'],
    'insurance': ['بیمه', 'نمایندگی بیمه', 'insurance agency', 'insurance broker'],
    'legal': ['دفتر حقوقی', 'وکالت', 'وکیل', 'موسسه حقوقی', 'law firm', 'lawyer', 'legal services'],
    'home_services': ['خدمات ساختمان', 'تعمیرات ساختمان', 'تعمیرات منزل', 'home services', 'handyman', 'home improvement'],
    'hvac': ['تاسیسات', 'تأسیسات', 'پکیج', 'کولر', 'آبگرمکن', 'hvac contractor', 'air conditioning repair service', 'heating contractor'],
    'carpet': ['قالیشویی', 'قالی شویی', 'carpet cleaning service'],
    'laundry': ['خشکشویی', 'خشک شویی', 'dry cleaner', 'laundry service'],
    'studio': ['آتلیه', 'استودیو عکاسی', 'عکاسی', 'photography studio', 'photographer'],
    'venue': ['تالار', 'باغ تالار', 'تشریفات مجالس', 'wedding venue', 'banquet hall', 'event venue'],
}


def _canon(value):
    return re.sub(r'\s+', ' ', str(value or '').translate(PERSIAN_DIGITS).replace('\u200c', ' ').strip().lower())


def _normalize_mobile(value):
    raw = re.sub(r'\D', '', str(value or '').translate(PERSIAN_DIGITS))
    if raw.startswith('0098'):
        raw = '0' + raw[4:]
    elif raw.startswith('98') and len(raw) == 12:
        raw = '0' + raw[2:]
    elif raw.startswith('9') and len(raw) == 10:
        raw = '0' + raw
    return raw if re.fullmatch(r'09\d{9}', raw) else ''


def _normalize_phone(value):
    raw = re.sub(r'\D', '', str(value or '').translate(PERSIAN_DIGITS))
    if raw.startswith('0098'):
        raw = '0' + raw[4:]
    elif raw.startswith('98') and len(raw) >= 11:
        raw = '0' + raw[2:]
    return raw[:15]


def _header_map(headers):
    normalized = {_canon(h): h for h in headers if _canon(h)}
    out = {}
    for target, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            key = _canon(alias)
            if key in normalized:
                out[target] = normalized[key]
                break
    return out


def _unique_headers(values):
    headers = []
    seen = {}
    for index, value in enumerate(values):
        base = str(value or '').strip()
        if not base:
            base = f'__column_{index + 1}'
        count = seen.get(base, 0)
        seen[base] = count + 1
        headers.append(base if count == 0 else f'{base}__{count + 1}')
    return headers


def _header_candidate(values):
    headers = _unique_headers(values)
    hm = _header_map(headers)
    return headers, hm, ('business_name' in hm and len(hm) >= 2)


def _scan_header(iterator, max_rows=100):
    for row_number, values in enumerate(iterator, start=1):
        if row_number > max_rows:
            break
        headers, hm, ok = _header_candidate(values)
        if ok:
            return row_number, headers, hm
    raise ValueError('هدر جدول پیدا نشد. فایل باید ستونی مثل Name / title / business_name یا «نام کسب‌وکار» داشته باشد.')


def _dict_rows(iterator, headers):
    width = len(headers)
    for values in iterator:
        values = list(values or [])
        if not any(str(v or '').strip() for v in values):
            continue
        yield {headers[i]: values[i] if i < len(values) else '' for i in range(width)}


def _rows_from_upload(file_storage):
    filename = (file_storage.filename or '').lower()
    file_storage.stream.seek(0)

    if filename.endswith('.xlsx'):
        wb = load_workbook(file_storage.stream, read_only=True, data_only=True)
        ws = wb.active
        iterator = ws.iter_rows(values_only=True)
        header_row, headers, _ = _scan_header(iterator)
        return _dict_rows(iterator, headers), header_row, headers

    wrapper = io.TextIOWrapper(file_storage.stream, encoding='utf-8-sig', errors='ignore', newline='')
    iterator = csv.reader(wrapper)
    header_row, headers, _ = _scan_header(iterator)
    return _dict_rows(iterator, headers), header_row, headers


def _cell(row, hm, key):
    header = hm.get(key)
    return row.get(header) if header else ''


def _first_url(value):
    text = str(value or '').strip()
    if not text:
        return ''
    for piece in re.split(r'[\r\n,\s]+', text):
        if piece.startswith(('http://', 'https://')):
            return piece
    return text.splitlines()[0].strip()


def _instagram_value(value):
    text = str(value or '').strip()
    if not text:
        return ''
    matches = re.findall(r'https?://(?:www\.)?instagram\.com/[^\s,]+', text, flags=re.I)
    return matches[0].rstrip('/,') if matches else text


def detect_vertical(category, name=''):
    hay = _canon(f'{category} {name}')
    order = ['realestate','beauty','barber','auto','aesthetic','dentist','gym','trainer','language','repair','parts','carwash','fashion','gold','furniture','cabinet','restaurant','pet','mobile','immigration','travel','insurance','legal','home_services','hvac','carpet','laundry','studio','venue','education']
    for key in order:
        if any(_canon(word) in hay for word in CATEGORY_KEYWORDS.get(key, [])):
            return key
    return ''


def register_data_ingest(app, db_path):
    def db():
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def admin_only(fn):
        @wraps(fn)
        def wrapped(*args, **kwargs):
            if not session.get('admin'):
                return redirect(url_for('login', next=request.path))
            return fn(*args, **kwargs)
        return wrapped

    @app.post('/admin/data/import-raw')
    @admin_only
    def import_raw_business_database():
        upload = request.files.get('file')
        default_vertical = (request.form.get('vertical') or '').strip()
        campaign = (request.form.get('campaign') or 'OUTBOUND-LIVE').strip()
        source = (request.form.get('source') or 'business_database').strip()

        if not upload:
            flash('فایل CSV/XLSX انتخاب نشده.', 'error')
            return redirect(url_for('campaigns_hub'))

        try:
            iterator, header_row, detected_headers = _rows_from_upload(upload)
            first = next(iterator, None)
        except Exception as exc:
            flash(f'فایل قابل خواندن نیست: {exc}', 'error')
            return redirect(url_for('campaigns_hub'))

        if not first:
            flash(f'فایل بعد از هدر ردیف {header_row} هیچ داده‌ای ندارد.', 'error')
            return redirect(url_for('campaigns_hub'))

        hm = _header_map(detected_headers)
        if 'business_name' not in hm:
            flash(f'هدر در ردیف {header_row} پیدا شد، اما ستون نام کسب‌وکار شناسایی نشد.', 'error')
            return redirect(url_for('campaigns_hub'))

        looks_like_google = bool(hm.get('place_id') or hm.get('maps_url') or hm.get('rating') or hm.get('latitude') or hm.get('longitude'))
        if source == 'business_database' and looks_like_google:
            source = 'google_maps'

        conn = db()
        imported = skipped = mobile_count = landline_only = 0
        auto_mapped = place_duplicates = duplicate_count = no_contact = 0

        try:
            conn.execute('''CREATE TABLE IF NOT EXISTS lead_external_ids (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                provider TEXT NOT NULL,
                external_id TEXT NOT NULL,
                lead_id INTEGER NOT NULL,
                created_at TEXT NOT NULL,
                UNIQUE(provider, external_id)
            )''')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_lead_external_ids_lead ON lead_external_ids(lead_id)')

            for row in chain([first], iterator):
                name = str(_cell(row, hm, 'business_name') or '').strip()
                category = str(_cell(row, hm, 'category') or '').strip()
                vertical = default_vertical or detect_vertical(category, name)
                if not name or vertical not in VERTICALS:
                    skipped += 1
                    continue
                if not default_vertical:
                    auto_mapped += 1

                generic_phone = _cell(row, hm, 'phone_any')
                raw_mobile = _cell(row, hm, 'mobile') or generic_phone
                raw_landline = _cell(row, hm, 'landline') or generic_phone
                mobile = _normalize_mobile(raw_mobile)
                landline = _normalize_phone(raw_landline)
                if mobile and landline == mobile:
                    landline = ''
                primary = mobile or landline
                if not primary:
                    no_contact += 1
                    skipped += 1
                    continue
                if mobile:
                    mobile_count += 1
                else:
                    landline_only += 1

                address = str(_cell(row, hm, 'address') or '').strip()
                city = str(_cell(row, hm, 'city') or '').strip()
                place_id = str(_cell(row, hm, 'place_id') or '').strip()

                if place_id and conn.execute('SELECT 1 FROM lead_external_ids WHERE provider=? AND external_id=? LIMIT 1', ('google_maps', place_id)).fetchone():
                    place_duplicates += 1
                    skipped += 1
                    continue

                duplicate = conn.execute('''SELECT id FROM leads
                    WHERE vertical=? AND (
                        phone=? OR
                        (business_name=? AND COALESCE(address,'')=? AND ?!='') OR
                        (business_name=? AND COALESCE(city,'')=? AND COALESCE(address,'')='' AND ?!='')
                    ) LIMIT 1''', (vertical, primary, name, address, address, name, city, city)).fetchone()
                if duplicate:
                    duplicate_count += 1
                    skipped += 1
                    if place_id:
                        conn.execute('INSERT OR IGNORE INTO lead_external_ids(provider,external_id,lead_id,created_at) VALUES(?,?,?,?)', ('google_maps', place_id, duplicate['id'], datetime.now().isoformat(timespec='seconds')))
                    continue

                instagram = _instagram_value(_cell(row, hm, 'instagram'))
                logo_url = _first_url(_cell(row, hm, 'logo_url'))
                meta = {
                    'campaign': campaign,
                    'variant': 'A' if imported % 2 == 0 else 'B',
                    'source': source,
                    'raw_category': category,
                    'owner': str(_cell(row, hm, 'owner') or '').strip(),
                    'mobile': mobile,
                    'landline': landline,
                    'sms_eligible': bool(mobile),
                    'website': str(_cell(row, hm, 'website') or '').strip(),
                    'google_maps_url': str(_cell(row, hm, 'maps_url') or '').strip(),
                    'place_id': place_id,
                    'rating': str(_cell(row, hm, 'rating') or '').strip(),
                    'reviews': str(_cell(row, hm, 'reviews') or '').strip(),
                    'latitude': str(_cell(row, hm, 'latitude') or '').strip(),
                    'longitude': str(_cell(row, hm, 'longitude') or '').strip(),
                    'opening_hours': str(_cell(row, hm, 'opening_hours') or '').strip(),
                    'business_status': str(_cell(row, hm, 'status') or '').strip(),
                    'import_header_row': header_row,
                }
                slugbase = re.sub(r'[^a-z0-9\u0600-\u06ff]+', '-', name.lower()).strip('-')[:42] or 'business'
                slug = f'{slugbase}-{secrets.token_hex(2)}'
                cur = conn.execute('''INSERT INTO leads(slug,business_name,vertical,phone,city,address,instagram,logo_url,accent,meta_json,status,created_at)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?)''', (
                    slug, name, vertical, primary, city, address, instagram, logo_url,
                    '#5b4df5', json.dumps(meta, ensure_ascii=False), 'new', datetime.now().isoformat(timespec='seconds')
                ))
                lead_id = cur.lastrowid
                if place_id:
                    conn.execute('INSERT OR IGNORE INTO lead_external_ids(provider,external_id,lead_id,created_at) VALUES(?,?,?,?)', ('google_maps', place_id, lead_id, datetime.now().isoformat(timespec='seconds')))
                imported += 1
                if imported % 1000 == 0:
                    conn.commit()

            conn.commit()
        except Exception as exc:
            conn.rollback()
            flash(f'ایمپورت متوقف شد و چیزی نیمه‌کاره ذخیره نشد. خطا: {type(exc).__name__}: {exc}', 'error')
            return redirect(url_for('campaigns_hub'))
        finally:
            conn.close()

        source_label = 'Google Maps' if source == 'google_maps' else source
        flash(
            f'هدر ردیف {header_row} شناسایی شد؛ {imported} لید از {source_label} وارد شد؛ '
            f'{mobile_count} موبایل برای SMS؛ {landline_only} فقط تلفن ثابت برای Call Desk؛ '
            f'{no_contact} بدون شماره؛ {duplicate_count} تکراری؛ {place_duplicates} Place ID تکراری؛ '
            f'{skipped} ردیف در مجموع رد شد؛ {auto_mapped} صنف خودکار تشخیص داده شد.',
            'success'
        )
        return redirect(url_for('campaigns_hub'))
